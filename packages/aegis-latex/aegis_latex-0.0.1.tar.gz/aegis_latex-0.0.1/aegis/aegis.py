#!/usr/bin/env python
# -*- coding: utf-8 -*-

# Copyright (c) 2020, Marcelo Lares
# License: MIT
#   Full text: https://github.com/mlares/aegis/blob/master/LICENSE

"""AEGIS: Academic Exam Generator for Interchange and Shuffe.

The purpose is to compose a collection of exams from a set of several
versions of a number of problems. This is useful to generate several
examns with a similar dificulty. It also can be used to generate
alternative exams from a pool of exercises.
"""
import jinja2
import itertools
import subprocess as sp
import os
import random


class Exam():
    """
    Exam (class): tools to generate exams with random exercises.

    Methods
    -------
    load_template : load template
    """

    def __init__(self):
        """Initialize an instance of class Exam.

        This class is used to generate many exams from a batch of
        exercises.
        """
        self.items = []
        self.subitems = []

    def load_template(self, template_file):
        """Load a template.

        Parameters
        ----------
        template_file: str
            Name of the file containing the latex template

        Returns
        -------
        Updates the self.template variable.
        """
        exists = os.path.exists(template_file)
        if exists:
            pass
        else:
            print('ERROR: template file not found or can not be loaded')
            self.template = None
            return None

        a = template_file.split('/')
        template_dir = '/'.join(a[:-1]) + '/'
        template_filename = a[-1]
        templateLoader = jinja2.FileSystemLoader(searchpath=template_dir)

        latex_jinja_env = jinja2.Environment(block_start_string=r"\BLOCK{",
                                             block_end_string='}',
                                             variable_start_string=r'\VAR{',
                                             variable_end_string='}',
                                             comment_start_string=r'\#{',
                                             comment_end_string='}',
                                             line_statement_prefix='%%',
                                             line_comment_prefix='%#',
                                             trim_blocks=True,
                                             autoescape=False,
                                             loader=templateLoader)

        # Make LaTeX file
        template = latex_jinja_env.get_template(template_filename)
        self.template = template

    def load_items(self, idir, items, subitems):
        """Load exercises to compile exams.

        Parameters
        ----------
        idir: str
           Directory where latex files are stored

        items: list
           List of numbers representing exercises

        subitems: list
           List of lists that contain the versions for each one of the
           exercises in "items".
        """
        exs = []
        k = -1
        for e in items:
            exs.append([])
            k += 1
            for v in subitems[k]:
                se = str(e).zfill(2)
                sv = str(v).zfill(2)
                fname = f"{idir}/e{se}_v{sv}.tex"
                with open(fname) as f:
                    txt = f.read()
                exs[k].append(txt)

        self.exs = exs

    def generate(self, N=0, output_dir='./',
                 shuffle=True, all_permutations=False,
                 makepdfs=False, interactive=False):
        """Generate exams suffling and randomly chosing items.

        Parameters
        ----------
        N : int (optional)
            The number of exams to generate.  If N is greater than the
            number of iterations, some exams will be repeated (by the
            Pigeon-hole theorem.)

        shuffle : boolean (optional)
            If True, shuffle the versions of the exercises to generate
            random versions of the exams. Dafault: True

        all_permutations : boolean (optional)
            If True, generate the complete list of possible combinations of the
            versions of the exercises. If N is present, it will be
            ignored. Default: False.

        makepdfs : boolean
            If True, compile PDF files from latex files. Default: False.

        interactive: boolean
            If True, return a list with the version used on the exams.

        Returns
        --------
        ex_list : list
            A list containing the versions of the exercises. Only
            returned if "interactive=True".

        """
        if all_permutations:
            shuffle = False

        if not isinstance(output_dir, str):
            print('output_dir not valid in function gen examples,')
            print('using output_dir=exams/')
            output_dir = "exams"

        if os.path.isdir(output_dir):
            msg = f"Saving PDF files in {output_dir}."
            print(msg)
        else:
            os.makedirs(output_dir)

        r = []
        nv = []
        for e in self.exs:
            i = len(e)
            r.append(range(i))
            nv.append(i)
        r = tuple(r)

        nn = 1
        for n in nv:
            nn = nn * n
        Ntot = nn

        it = itertools.product(*r)

        if shuffle:

            if N == 0:
                N = Ntot

            comb = []
            for x in it:
                comb.append(x)

            ex_list = []
            for c in range(N):
                ch = random.sample(range(Ntot), 1)[0]
                versions = list(comb[ch])
                ex_list.append(versions)
                exs = []
                for j, k in enumerate(versions):
                    exs.append(self.exs[j][k])

                texname = 'parcial_' + str(c + 1).zfill(4) + '.tex'
                texfile = '/'.join([output_dir, texname])
                print(f'Generando parcial en archivo {texfile}')

                target = open(texfile, 'w')
                target.write(self.template.render(exs=exs))
                target.close()

                if makepdfs:
                    os.popen(f'cp famaf.cls {output_dir}/')
                    os.popen(f'cp logo_famaf.png {output_dir}/')

                    cmd = ['pdflatex', '-interaction', 'nonstopmode', texname]
                    print(cmd)
                    source_dir = os.getcwd()
                    os.chdir(output_dir)
                    for _ in range(3):
                        proc = sp.Popen(cmd, stdout=sp.PIPE)
                        proc.communicate()
                    os.chdir(source_dir)

        if all_permutations:

            ex_list = []
            for c, v in enumerate(it):

                exs = []
                for i, k in enumerate(list(v)):
                    exs.append(self.exs[i][k])
                ex_list.append([c, k])

                texname = 'parcial_' + str(c + 1).zfill(4) + '.tex'
                texfile = '/'.join([output_dir, texname])
                print(f'Generando parcial en archivo {texfile}')

                target = open(texfile, 'w')
                target.write(self.template.render(exs=exs))
                target.close()

                if makepdfs:
                    source_dir = os.getcwd()
                    os.chdir(output_dir)
                    cmd = ['pdflatex', '-interaction', 'nonstopmode', texname]
                    for _ in range(3):
                        proc = sp.Popen(cmd, stdout=sp.PIPE)
                        proc.communicate()

                    os.chdir(source_dir)
        self.ex_list = ex_list
        if interactive:
            return ex_list

    def gen_excell(self, output_dir='./'):
        """Generate an Excell file with the contents of the exams.

        Parameters
        ----------
        output_dir: directory where to put the excell file

        Returns
        --------
        None
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "exam"

        Aord = ord('A')
        for i in range(len(self.exs)):
            ws[f'{chr(Aord + i + 1)}1'] = f"Ex. {i+1}"

        for i, e in enumerate(self.ex_list):
            ws[f"{chr(Aord)}{i+2}"] = f"examen {i+1}"
            for j, v in enumerate(e):
                print(f"{chr(Aord+j)}, {i+2} : {v+1} ")
                ws[f"{chr(Aord + j + 1)}{i + 2}"] = v + 1

        wb.save('exams_versions.xlsx')


def gen_examples(N_problems=1, N_versions=[[1]],
                 dir_exams='exams/'):
    """Generate exams example files.

    Parameters
    ----------
    N_problems: int
        Numbers of the problems in the exams
    N_versions: list of lists
        Numbers of the versions to be used in the problems
    dir_exams: str
        Directory where latex files are stored

    Returns
    -------
    problems: list
        List of the numbers of the problems
    versions: list of lists
        Lists with the numbers of the versions
    """
    from os import path, makedirs

    verbose = True
    if not isinstance(dir_exams, str):
        print('dir_exams not valid in function gen examples,')
        print('using dir_exams=exams/')
        dir_exams = "exams"
    if not path.isdir(dir_exams):
        print(f"Directory {dir_exams} does not exist")
        makedirs(dir_exams)
        if verbose:
            print("Directory ", dir_exams, " Created ")

    for ip, p in enumerate(range(N_problems)):
        for iv, v in enumerate(N_versions[ip]):

            se = str(p + 1).zfill(2)
            sv = str(v).zfill(2)
            filename = f"e{se}_v{sv}.tex"
            filename = '/'.join([dir_exams, filename])
            content = (f"This is the problem number {ip+1}, "
                       f"version {iv+1}.")

            file = open(filename, 'w')
            file.write(content)
            file.close()

    problems = list(range(N_problems))
    versions = []
    for p in problems:
        vs = N_versions[p]
        versions.append(vs)
    problems = [p + 1 for p in problems]

    return problems, versions

