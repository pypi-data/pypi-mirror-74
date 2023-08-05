from openpyxl import Workbook
from openpyxl import load_workbook

FILE_PATH = 'people.xlsx'

dictionary = {
        'id': 0, 0: 'id',
        'position_id': 1, 1: 'position_id',
        'first_name': 2, 2: 'first_name',
        'middle_name': 3, 3: 'middle_name',
        'last_name': 4, 4: 'last_name',
        'bday': 5, 5: 'bday',
        'bmonth': 6, 6: 'bmonth',
        'byear': 7, 7: 'byear',
        'image_link': 8, 8: 'image_link',
        'link': 9, 9: 'link'
    }


def save_excel(path, people):
    try:
        wb = load_workbook(path+FILE_PATH)
        ws = wb.active
        i = 1
        while ws.cell(row=i, column=1).value != None:
            i = i + 1
        if i - 1 < 0:
            print("Error params")
        else:
            count = 0
            for col in ws.iter_cols(min_row=i, max_col=10, max_row=i + len(people)):
                for j in range(0, len(people)):
                    if count == 0:
                        col[j].value = i + j - 1
                    else:
                        col[j].value = people[j][dictionary[count]]
                count = count + 1
        wb.save(path+FILE_PATH)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = "People"
        count = 0
        for col in ws.iter_cols(min_row=1, max_col=10, max_row=len(people) + 1):
            for i in range(0, len(people) + 1):
                if i == 0:
                    col[i].value = dictionary[count]
                else:
                    if count == 0:
                        col[i].value = i
                    else:
                        col[i].value = people[i - 1][dictionary[count]]
            count = count + 1
        wb.save(path+FILE_PATH)
    return 0
