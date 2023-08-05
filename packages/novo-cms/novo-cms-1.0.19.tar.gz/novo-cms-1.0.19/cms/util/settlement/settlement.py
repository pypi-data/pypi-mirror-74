# -*- encoding: utf8 -*-
"""
    CMS SubProject Crawler
"""
import os
import re
import sys
import time
import json
import math
import logging
import textwrap
import datetime

from collections import defaultdict

try:
    import cPickle as pickle
except ImportError:
    import pickle

import openpyxl

import requests

from .title import TITLE
from . import parse_report_data, write_title, write_sheet2

from cms.util import MyLogger, User


class Settlement(object):
    def __init__(self, report_files, outfile, lims_file=None, logger=None, filter_report=False):
        self.report_files = report_files
        self.outfile = outfile
        self.lims_file = lims_file
        self.filter_report = filter_report
        self.logger = logger or MyLogger()

    def get_stage_data(self):

        self.logger.info('parsing report data: {} ...'.format(self.report_files))

        return parse_report_data(self.report_files, filter_report=self.filter_report)

    def get_subproject_ids(self, stage_data):

        dup_filter = {}
        with open(self.outfile, 'w') as out:
            for stagecode in stage_data:
                context = stage_data[stagecode]['result']
                subprojectnum = context['subprojectnum']
                if subprojectnum in dup_filter:
                    continue
                dup_filter[subprojectnum] = 1
                out.write('{}\n'.format(subprojectnum))

        self.logger.info('save file: {}'.format(self.outfile))

    def get_lims_data(self):
        lims_data = {}

        if self.lims_file and os.path.isfile(self.lims_file):
            self.logger.info('loading lims data: {} ...'.format(self.lims_file))
            with open(self.lims_file) as f:
                lims_data = pickle.load(f)

            lims_data = {k2:v2 for k,v in lims_data.items() for k2,v2 in v.items()}

        return lims_data

    def create_book(self, stage_data):
        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = unicode('子项目分期结算')
        sheet2 = book.create_sheet(unicode('过滤后结算报表'))
        write_title(sheet1, TITLE)

        fields = [each[1] for part in TITLE for each in part]

        lims_data = self.get_lims_data()

        row = 1
        for stagecode in stage_data:
            
            context = stage_data[stagecode]['result']
            # print context;exit()
            context['jobname'] = '分期数据结题'

            row += 1

            # LIMS数据补充
            lims_info = lims_data.get(stagecode)
            if lims_info:
                context['path_date'] = lims_info['path_date']
                context['xiadan_date'] = lims_info['xiadan_date']

                # 信息计划完成时间：
                #   - 默认加6天
                #   - 样本数大于48，每超出24个样本再加1天
                if context['path_date']:
                    days = 6
                    if context['samplenum'] and context['samplenum'] > 48:
                        days += math.ceil((context['samplenum'] - 48) / 24.)
                    context['bi_plan_time'] = context['path_date'] + datetime.timedelta(days=days)
            
            for col, key in enumerate(fields, 1):
                value = context.get(key)

                value = value or ''

                if isinstance(value, datetime.datetime):
                    value = value.strftime('%Y-%m-%d')

                sheet1.cell(row=row, column=col, value=value)

        write_sheet2(sheet2, stage_data)

        book.save(self.outfile)
        self.logger.info('\033[33msave file: {}\033[0m'.format(self.outfile))


def main(**args):

    start_time = time.time()

    logger = MyLogger('Settlement', **args)

    settlement = Settlement(args['report'], args['out'], lims_file=args['lims_data'], filter_report=args['filter_report'], logger=logger)

    stage_data = settlement.get_stage_data()

    if args['get_subproject']:
        settlement.get_subproject_ids(stage_data)
    else:
        settlement.create_book(stage_data)

    logger.info('\033[32mtime used: {:.1f}s\033[0m'.format(time.time() - start_time))


def parser_add_settlement(parser):

    parser.description = __doc__
    parser.epilog = textwrap.dedent('''
        \033[36mexamples:
            %(prog)s -r yewuxian.xlsx
        \033[0m
    ''')

    parser.add_argument(
        '-o', '--out', help='the output filename [default: %(default)s]', default='settlement.xlsx')

    parser.add_argument(
        '-r', '--report', help='the report excel which contains sample informations', nargs='+', required=True)

    parser.add_argument('-gs', '--get-subproject', help='get the subproject ids from report', action='store_true')

    parser.add_argument('-f', '--filter-report', help='filter the report with settlement <= 0', action='store_true')

    parser.add_argument('-lims', '--lims-data', help='the lims data to update stage information')

    parser.set_defaults(func=main)
