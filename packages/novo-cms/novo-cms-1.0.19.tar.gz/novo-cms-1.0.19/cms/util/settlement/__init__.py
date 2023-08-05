# -*- encoding: utf8 -*-
import re
import sys
import json
import datetime
import warnings
from collections import defaultdict
from copy import deepcopy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side


warnings.simplefilter("ignore")


def write_title(sheet1, title):

    colors = '64F336 EBBF1D 00FFF3'.split()
    col = 0
    for part, color in zip(title, colors):
        for field in part:
            value = field[0]
            col += 1
            _ = sheet1.cell(row=1, column=col, value=value)
            _.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            _.font = Font(bold=True)

            sheet1.column_dimensions[get_column_letter(col)].width = 20


def write_sheet2(sheet, stage_data):
    title_done = False

    for index, stagecode in enumerate(stage_data, 2):
        title = stage_data[stagecode]['title']
        if not title_done:
            for column, value in enumerate(title, 1):
                _ = sheet.cell(row=1, column=column, value=value)
                _.font = Font(bold=True)
                color = '64F336'
                _.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet.column_dimensions[get_column_letter(column)].width = 20
            title_done = True
        row = stage_data[stagecode]['rows']
        row[0] = index - 1
        for column, value in enumerate(row, 1):
            sheet.cell(row=index, column=column, value=value)


def parse_report_data(report_xlsx_list, filter_report=False):
    """
        业务线维度 - 结算报表信息提取
    """
    fields = [
        ['子项目编号', 'subprojectnum'],
        ['子项目名称', 'subprojectname'],
        ['子项目产品类型', 'productname'],
        ['运营经理', 'pc_name'],
        ['信息负责人', 'bi_name'],
        ['部门', 'department'],
        ['系统分期编号', 'stagecode'],
        ['结算样本个数', 'samplenum'],
        ['理论结题时间', 'pc_plan_time'],
        ['实际结题时间', 'pc_complete_time'],
        ['理论结题时间', 'pc_plan_time'],

        ['结算金额', 'settlement'],
    ]

    stage_data = defaultdict(dict)

    for report_xlsx in report_xlsx_list:
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.get_active_sheet()
        for row in ws.rows:
            if row[0].value == '序号':
                title = [each.value.encode() for each in row if each.value]
                print 'title: ', '\t'.join(title)
                continue

            idx = row[0].value
            try:
                idx = float(idx)
            except:
                pass

            if not isinstance(idx, float):
                continue

            linelist = [each.value for each in row]

            context = dict(zip(title, linelist))

            result = {k2: context[k1] for k1, k2 in fields}

            samplenum = result['samplenum']
            if samplenum:
                result['samplenum'] = int(samplenum)

            stagecode = result['stagecode']

            if stagecode not in stage_data:
                stage_data[stagecode]['title'] = title
                stage_data[stagecode]['rows'] = []
                stage_data[stagecode]['result'] = []
                stage_data[stagecode]['settlement'] = []

            stage_data[stagecode]['rows'].append(linelist)
            stage_data[stagecode]['result'].append(result)
            stage_data[stagecode]['settlement'].append(result['settlement'])

            # print stage_data;exit()

    # 根据结算金额总和和个数进行过滤
    #   - 总和为0，说明正好冲销掉了，删除该分期
    #   - 总和不为0，个数大于1，说明为多次结算，保留最后一个，替换结算金额
    res_data = defaultdict(dict)

    for code in stage_data:
        settlement = stage_data[code]['settlement']
        title = stage_data[code]['title']

        if (sum(settlement) <= 0) and filter_report:
            print 'filtered as settlement<=0 for stagecode: {}'.format(code)
            continue

        rows = stage_data[code]['rows'][-1]
        result = stage_data[code]['result'][-1]

        if len(settlement) > 1:
            rows[title.index('结算金额')] = sum(settlement)
            result['settlement'] = sum(settlement)

        res_data[code]['title'] = title
        res_data[code]['rows'] = rows
        res_data[code]['result'] = result

    return dict(res_data)
