# -*- encoding: utf8 -*-
import re
import sys
import warnings
from collections import defaultdict
from copy import deepcopy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Color, colors, Alignment, Border, Side


warnings.simplefilter("ignore")


def write_title(sheet1, sheet2, title, stage_title, stage_max=5):

    stage_titles = []
    for i in range(1, stage_max+1):
        tmp_list = deepcopy(stage_title)
        tmp_list[0][0] = '任务{} 任务名称'.format(i)
        stage_titles += tmp_list

    titles = title + [stage_titles]

    colors = '64F336 EBBF1D 00FFF3 3184EA'.split()
    col = 0
    for part, color in zip(titles, colors):
        for field in part:
            if isinstance(field, list):
                value = field[0]
            else:
                value = field
            col += 1
            _ = sheet1.cell(row=1, column=col, value=value)
            _.fill = PatternFill(start_color=color,
                                 end_color=color,
                                 fill_type="solid")
            _.font = Font(bold=True)
            if col > 1:
                sheet1.column_dimensions[get_column_letter(col)].width = 20

    color = 'FF8000'
    for col, value in enumerate(['合同编号', '子项目编号', '工序名称', '工序类型', '样本个数', '数据量'], 1):
        _ = sheet2.cell(row=1, column=col, value=value)
        _.font = Font(bold=True)
        _.fill = PatternFill(start_color=color,
                             end_color=color,
                             fill_type="solid")
        sheet2.column_dimensions[get_column_letter(col)].width = 24


def parse_report_data(report_xlsx_list):
    """
        从报表中获取结算样本信息
            业务线维度 - 结算样本个数，结算金额
            合同维度  - 样本数，数据量，产品类型
    """
    report_data = {}

    if not report_xlsx_list:
        return report_data

    for report_xlsx in report_xlsx_list:
        wb = openpyxl.load_workbook(report_xlsx)
        ws = wb.get_active_sheet()
        for row in ws.rows:
            if row[0].value == '序号':
                title = [each.value.decode() for each in row]
                if '子项目编号' in title:
                    key_indexes = [title.index('子项目编号')]
                    value_indexes = map(title.index, ['结算样本个数', '结算金额'])
                else:
                    key_indexes = map(title.index, ['项目编号', '产品编码'])
                    value_indexes = map(title.index, ['样本数', '数据量', '产品类型'])
                continue
            elif not isinstance(row[0].value, float):
                continue

            keys = tuple(each.value for each in map(
                row.__getitem__, key_indexes))
            if len(keys) == 1:
                keys = keys[0]

            values = [each.value for each in map(
                row.__getitem__, value_indexes)]

            if keys not in report_data:
                report_data[keys] = defaultdict(int)

            if '子项目编号' in title:
                if values[0] and values[0].isdigit():
                    report_data[keys]['js_samplenum'] += int(values[0])
                if values[1]:
                    report_data[keys]['js_money'] += float(values[1])
            else:
                if values[0] and values[0].isdigit():
                    report_data[keys]['samplenum'] += int(values[0])
                if values[1]:
                    report_data[keys]['datasize'] += int(values[1])
                if values[2]:
                    report_data[keys]['processtypename'] = values[2]

    return report_data


def parse_config(configfile):
    wb = openpyxl.load_workbook(configfile)
    ws = wb.get_active_sheet()

    config = {
        'common': {},
        'product': {},
        'sop': {},
    }
    for row in ws.rows:

        if re.match(r'#|一级|工序名称', str(row[0].value)) or not row[0].value:
            continue

        if not row[1].value:
            config['common'][row[0].value] = row[2].value
        elif isinstance(row[1].value, long):
            config['product'][row[0].value] = [
                row[1].value, row[2].value, row[3].value]
        else:
            config['sop']['__'.join(
                [row[0].value, row[1].value])] = row[2].value

    return config
