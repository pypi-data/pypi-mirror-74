# -*- encoding: utf8 -*-
"""
    CMS SubProject Crawler
"""
import os
import re
import sys
import time
import json
import math
import logging
import textwrap
import codecs

from collections import defaultdict

try:
    import cPickle as pickle
except ImportError:
    import pickle

import openpyxl

import requests

from .title import TITLE, STAGE_TITLE
from . import parse_config, parse_report_data, write_title

from cms.util import MyLogger, User

ROOT_DIR = os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.realpath(__file__))))

DEFAULT_CONFIGS = [
    './cms.config.xlsx',
    os.path.join(os.path.expanduser('~'), 'cms.config.xlsx'),
    os.path.join(ROOT_DIR, 'conf', 'cms.config.xlsx'),
]

for each in DEFAULT_CONFIGS:
    if os.path.isfile(each):
        DEFAULT_CONFIG = each
        break

reload(sys)
sys.setdefaultencoding('utf-8')


class CMS(object):
    base_url = 'http://cms.novogene.com'

    def __init__(self, username, password, logger=None):
        self.username = username
        self.password = password
        self.logger = logger or MyLogger()
        self.session = requests.session()

    def login(self):
        url = self.base_url + '/core/login/login!login.action'
        formdata = {
            'loginInfo.usercode': self.username,
            'loginInfo.userpass': self.password,
        }
        resp = self.session.post(url, data=formdata)
        loginMessage = resp.json()['loginMessage']
        if loginMessage == 'success':
            self.logger.info('Login successfully!')
            return True
        self.logger.error('Login failed as: {}'.format(loginMessage))
        return False

    @property
    def urls(self):
        urls = {
            'selectProjectInfosByCond': '/nhzy/project/project!selectProjectInfosByCond.action',
            'selectSubprojectInfosByCond': '/nhzy/subproject/subproject!selectSubprojectInfosByCond.action',
            'selectStageInfosByCond': '/nhzy/settlementbill/stage!selectStageInfosByCond.action',
            'selectContractInfoByContractsno': '/crm/contract/contract!selectContractInfoByContractsno.action',
            'selectQuotationproductInfosByCond': '/nhzy/projectquotation/quotationproduct!selectQuotationproductInfosByCond.action',
            'selectQuoprocessInfosByCond': '/nhzy/projectquotation/quoprocess!selectQuoprocessInfosByCond.action',
        }
        return urls

    def selectProjectInfosByCond(self, subprojectoperatorcode=None, projectnum=None):
        """
            项目列表
        """
        url = self.base_url + self.urls['selectProjectInfosByCond']

        formdata = {
            # 'cond.kftype': 'A',
            'cond.subprojectoperatorcode': subprojectoperatorcode,
            'cond.projectnum': projectnum,
            'limit': 99999999,
        }

        resp = self.session.post(url, data=formdata).json()
        return resp['projectInfos']

    def selectSubprojectInfosByCond(self, subprojectoperatorcode=None, projectnum=None, subprojectnum=None, profitcodes=None):
        """
            子项目列表
        """
        url = self.base_url + self.urls['selectSubprojectInfosByCond']
        formdata = {
            'cond.batchsubprojectnum': 1,
            'cond.subprojectoperatorcode': subprojectoperatorcode,
            'cond.projectnum': projectnum,
            'cond.subprojectnum': subprojectnum,
            'limit': 99999999,
        }

        subprojectInfos = []
        if profitcodes:
            for profitcode in profitcodes:
                formdata['cond.profitcode'] = profitcode
                subprojectInfos += self.session.post(url, data=formdata).json()[
                    'subprojectInfos']
        else:
            subprojectInfos = self.session.post(url, data=formdata).json()[
                'subprojectInfos']

        return subprojectInfos

    def selectStageInfosByCond(self, subprojectnum):
        """
            分期查询
        """
        url = self.base_url + self.urls['selectStageInfosByCond']
        formdata = {
            'cond.subprojectcode': subprojectnum,
            'limit': 99999999,
        }
        resp = self.session.post(url, data=formdata).json()

        return resp['stageInfos']

    def selectContractInfoByContractsno(self, contractno):
        """
            合同查询
        """
        url = self.base_url + self.urls['selectContractInfoByContractsno']
        formdata = {
            'cond.contractsno_equals': contractno,         # this is much faster than 'cond.contractsno'
        }
        resp = self.session.post(url, data=formdata).json()

        if len(resp['contractInfo']) != 1:
            self.logger.warn(
                'not unique contract number: {}'.format(contractno))
            exit(1)

        return resp['contractInfo']

    def selectQuotationproductInfosByCond(self, quotationid, pcode=None):
        """
            合同报价查询
        """
        url = self.base_url + self.urls['selectQuotationproductInfosByCond']
        formdata = {
            'cond.kfquotationid': quotationid,
            'cond.pcode': pcode,
            'limit': 99999999,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quotationproductInfos']

    def selectQuoprocessInfosByCond(self, kfquotationproductid):
        """
            工序查询
        """
        url = self.base_url + self.urls['selectQuoprocessInfosByCond']
        formdata = {
            'cond.kfquotationproductid': kfquotationproductid,
            'limit': 99999999,
        }
        resp = self.session.get(url, params=formdata).json()

        return resp['quoprocessInfos']

    def get_quotation_data(self, contractno, pcode):
        quotation_data = defaultdict(list)

        contract = self.selectContractInfoByContractsno(contractno)[0]
        quotation_products = self.selectQuotationproductInfosByCond(contract['quotationid'], pcode)

        # 上机数据量，样本数
        quotation_data['samplenum'] = 0
        quotation_data['totalproductdata'] = 0
        
        for quotation in quotation_products:
    

            # 工序列表
            quoprocesses = self.selectQuoprocessInfosByCond(
                quotation['kfquotationproductid'])
            for quoprocess in quoprocesses:

                if '上机' in quoprocess['processcode']:
                    quotation_data['totalproductdata'] += float(
                        quoprocess['datasize']) * int(quoprocess['samnum'])

                    quotation_data['samplenum'] += int(quoprocess['samnum'])

                if quoprocess['datasize'] and quoprocess['datasize'].isdigit():
                    quoprocess['datasize'] = int(quoprocess['datasize'])
                else:
                    quoprocess['datasize'] = ''

                quotation_data['processes'] += [
                    {
                        'processcode': quoprocess['processcode'],
                        'samnum': int(quoprocess['samnum']),
                        'datasize': quoprocess['datasize'],
                        'processtypename': quoprocess['processtypename']
                    }
                ]

            quotation_data['processtypename'] += quotation['processtypename'].split(',')

        return dict(quotation_data)

    def get_sub_infos(self, sub_projects, limit=None, subproject=None, project=None):

        if limit:
            return sub_projects[:limit]
        
        for obj, key in zip([subproject, project], ['subprojectnum', 'projectnum']):
            if obj:
                if os.path.isfile(obj):
                    with open(obj) as f:
                        key_list = [line.strip() for line in f]
                else:
                    key_list = obj.split(',')
                return filter(lambda x: x[key] in key_list, sub_projects)
    
        return sub_projects

    def parse_data(self, config_file, report_file, lims_file):

        # 配置文件解析
        self.logger.info('\033[32mparsing config data: {} ...\033[0m'.format(config_file))
        config_data = parse_config(config_file)
        self.logger.debug('\033[32mConfig Data: \n{}\033[0m'.format(
            json.dumps(config_data, ensure_ascii=False, indent=2)))

        # 报表文件解析
        if report_file:
            self.logger.info('\033[32mparsing report data: {} ...\033[0m'.format(', '.join(report_file)))
        report_data = parse_report_data(report_file)

        # LIMS数据读取
        lims_data = {}
        if lims_file and os.path.isfile(lims_file):
            self.logger.info('\033[32mparsing lims data: {} ...\033[0m'.format(lims_file))
            lims_data = pickle.load(open(lims_file))
            self.logger.info('found {} subprojects from {}'.format(len(lims_data), lims_file))

        return config_data, report_data, lims_data

    def deal_processes(self, sub_info, sheet2, indexes, config_data, sop_not_in_config, product_not_in_config):
        # SOP明细 -- 生产成本，不考虑分析
        # 测序前SOP成本 = SUM(样本数 * 对应单价)
        # 生产SOP成本 = 测序前SOP成本 + 数据量 * 单价 * pooling系数

        # 获取config中产品单价信息
        product_info = config_data['product'].get(unicode('{pname}'.format(**sub_info)))
        if not product_info:
            product_not_in_config.add((sub_info['subprojectnum'], sub_info['pname']))
            ave_sam_product, pooling, sop_price = 1, 1, 1
        else:
            ave_sam_product, pooling, sop_price = product_info

        if sub_info['samplenum']:
            sub_info['pc_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(indexes['sheet1'], ave_sam_product, sub_info['pc_ave_time'])
            sub_info['bi_time'] = '=ROUNDUP(J{}/{},0) * {}'.format(indexes['sheet1'], ave_sam_product, sub_info['bi_ave_time'])
            sub_info['pc_bi_time'] = '=W{0}+X{0}'.format(indexes['sheet1'])

        if sub_info.get('processes'):
            sopcost_before = []
            sopcost_production = []
            sop_filter = {}
            for process in sub_info['processes']:
                sop_key = '{processcode}__{processtypename}'.format(**process)

                # 工序去重
                if process['processcode'] in sop_filter:
                    continue
                sop_filter[process['processcode']] = 1

                indexes['sheet2'] += 1

                ave_sample = config_data['sop'].get(unicode(sop_key))

                if ave_sample:
                    self.logger.debug('{sop_key}: {ave_sample}'.format(**locals()))

                    if '上机' in process['processcode']:  # 测序成本
                        sopcost_production += ['K{}*{}*{}'.format(indexes['sheet1'], ave_sample, pooling)]
                    elif any([each in process['processcode'] for each in ('提取', '库检', '检测', '建库')]):  # 测序前生产成本
                        sopcost_before += ['J{}*{}'.format(indexes['sheet1'], ave_sample)]
                elif '分析' not in process['processcode']:
                    sop_not_in_config.add(
                        (sub_info['subprojectnum'], process['processcode'], process['processtypename']))

                sheet2.cell(row=indexes['sheet2'], column=1, value=sub_info['contractno'])
                sheet2.cell(row=indexes['sheet2'], column=2, value=sub_info['subprojectnum'])
                sheet2.cell(row=indexes['sheet2'], column=3, value=process['processcode'])
                sheet2.cell(row=indexes['sheet2'], column=4, value=process['processtypename'])
                sheet2.cell(row=indexes['sheet2'], column=5, value=process['samnum'])
                sheet2.cell(row=indexes['sheet2'], column=6, value=process['datasize'])

            if sopcost_before:
                sub_info['sopcost_before'] = '=' + ' + '.join(sopcost_before)
            else:
                sub_info['sopcost_before'] = 0

            if sopcost_production:
                sub_info['sopcost_production'] = '=L{} + '.format(indexes['sheet1']) + ' + '.join(sopcost_production)
            else:
                sub_info['sopcost_production'] = 0
        else:  # 对于没有sop的项目, 直接用单价*样本数?
            sub_info['sopcost_before'] = '=J{i}*{price}'.format(i=indexes['sheet1'], price=sop_price)
            sub_info['sopcost_production'] = '=J{i}*{price}'.format(i=indexes['sheet1'], price=sop_price)

    def deal_stages(self, sub_info, lims_data, stage_fields):

        stages = self.selectStageInfosByCond(sub_info['subprojectnum'])

        stage_list = []

        for stage_info in stages:

            tmp_dict = dict(sub_info, **stage_info)

            stage_code = stage_info['installmentcode']

            if lims_data.get(stage_code):
                if 'path_date' in lims_data[stage_code]:
                    tmp_dict['path_date'] = lims_data[stage_code]['path_date'].strftime('%Y-%m-%d')

            stage_list += [tmp_dict.get(k, '') for k in stage_fields]

        return stage_list, len(stages)

    def deal_sub_infos(self, sub_infos, **kwargs):

        # 数据解析
        config_data, report_data, lims_data = self.parse_data(kwargs['config'], kwargs['report_file'], kwargs['lims_file'])

        # 表头处理
        fields = [each[1] for part in TITLE for each in part]
        stage_fields = [each[1] for each in STAGE_TITLE]

        book = openpyxl.Workbook()
        sheet1 = book.active
        sheet1.title = unicode('子项目列表')
        sheet2 = book.create_sheet(unicode('SOP明细'))

        # 最大分期数
        stage_max = 0

        total = len(sub_infos)
        indexes = {'sheet1': 1, 'sheet2': 1}

        # 记录配置文件中没有的工序和产品名称
        sop_not_in_config = set()
        product_not_in_config = set()

        for n, sub_info in enumerate(sub_infos, 1):
            percent = n * 100. / total
            if kwargs.get('progress'):
                sys.stderr.write('\033[K[{n}/{total}] \033[36m{percent:.1f}% \033[0mcompleted\r'.format(**locals()))
                sys.stderr.flush()

            self.logger.debug('[{n}/{total}] dealing with: {subprojectnum}'.format(n=n, total=total, **sub_info))

            sub_info['index'] = indexes['sheet1']
            indexes['sheet1'] += 1

            # 更新config data
            sub_info.update(config_data['common'])
            if kwargs.get('PM'):
                sub_info['PM'] = kwargs['PM']

            # 项目税前收入 => 剩余额度
            subprojecttotallimit = float(sub_info['subprojecttotallimit'])  # 子项目总额度
            settlementmoney = float(sub_info['settlementmoney'])            # 已结算金额
            subprojectleftlimit = subprojecttotallimit - settlementmoney    # 剩余额度
            sub_info['subprojectleftlimit'] = subprojectleftlimit


            # 项目编号 => 合同编号 => 工序列表 => 数据量，样本数
            try:
                project_info = self.selectProjectInfosByCond(projectnum=sub_info['projectnum'])[0]
            except:
                print sub_info['projectnum'], sub_info['subprojectnum']
                exit(1)

            sub_info['contractno'] = project_info['contractno']
            quotation_data = self.get_quotation_data(sub_info['contractno'], sub_info['pcode'])

            # 更新：样本数，总数据量
            sub_info.update(quotation_data)

            # 使用报表数据进行填充
            if report_data.get(sub_info['subprojectnum']):                # 业务线维度结算报表
                sub_info['js_samplenum'] = report_data[sub_info['subprojectnum']]['js_samplenum']
                sub_info['js_money'] = report_data[sub_info['subprojectnum']]['js_money']

            if report_data.get((sub_info['projectnum'], sub_info['pcode'])):  # 合同维度结算报表
                result = report_data[(sub_info['projectnum'], sub_info['pcode'])]

                if sub_info.get('totalproductdata') is None:
                    sub_info['totalproductdata'] = result['datasize']

                if sub_info.get('samplenum') is None:
                    sub_info['samplenum'] = result['samplenum']

                if sub_info.get('pname') is None:
                    sub_info['pname'] = result['processtypename']

            # 数据量，样本数 * (剩余额度 / 子项目总额度)
            if not subprojecttotallimit:
                self.logger.debug('{subprojectnum}的总额度为0'.format(**sub_info))
                limit_percent = 0
            else:
                limit_percent = float(subprojectleftlimit) / subprojecttotallimit

            if sub_info.get('samplenum'):
                sub_info['samplenum'] = int(math.ceil(int(sub_info['samplenum']) * limit_percent))

            if sub_info.get('totalproductdata'):
                sub_info['totalproductdata'] = int(math.ceil(int(sub_info['totalproductdata']) * limit_percent))

            # 处理工序数据
            self.deal_processes(sub_info, sheet2, indexes, config_data, sop_not_in_config, product_not_in_config)

            # 其他信息填充
            sub_info['compute_cost'] = '=K{i}*{gcluster}'.format(i=indexes['sheet1'], **sub_info)
            sub_info['pc_bi_cost'] = '=W{i}*{pccost} + X{i}*{bicost}'.format(i=indexes['sheet1'], **sub_info)

            sub_info['people_cost'] = '=I{i}/1.06*{apportion_people}'.format(i=indexes['sheet1'], **sub_info)
            sub_info['three_cost'] = '=I{i}/1.06*{apportion_three}'.format(i=indexes['sheet1'], **sub_info)

            sub_info['profit'] = '=I{i}/1.06 - M{i} - N{i} - P{i} -Q{i} - R{i}'.format(i=indexes['sheet1'])

            if subprojectleftlimit:
                sub_info['profit_rate'] = '=S{i}/I{i}'.format(i=indexes['sheet1'])

            # 填入数据
            linelist = [sub_info.get(k, '') for k in fields]

            stage_list, stage_number = self.deal_stages(sub_info, lims_data, stage_fields)
            stage_max = stage_number if stage_number > stage_max else stage_max

            # 写入当前行
            linelist += stage_list
            for col, value in enumerate(linelist, 1):
                try:
                    value = float(value)
                except:
                    try:
                        value = int(value)
                    except:
                        pass
                if isinstance(value, list):
                    value = ','.join(value)
                sheet1.cell(row=indexes['sheet1'], column=col, value=value)

        # 写入表头
        write_title(sheet1, sheet2, TITLE, STAGE_TITLE, stage_max=stage_max)
        book.save(kwargs['out'])


def main(**args):
    start_time = time.time()

    logger = MyLogger('CMS', **args)

    # 输入参数显示
    msg = []
    for k, v in args.iteritems():
        if k not in ('func', 'subparser_name'):
            if isinstance(v, list):
                v = ', '.join(v)
            msg += ['{:>10}:\t{}'.format(k, v)]
    logger.info('input arguments:\n\033[33m{}\033[0m'.format('\n'.join(msg)))

    # 登录用户名密码
    user = User(**args)
    username, password = user.get_user_pass()
    cms = CMS(username, password, logger=logger)
    if cms.login():
        user.save(username, password)
    else:
        exit(1)

    subprojectoperatorcode = args['subprojectoperatorcode']
    profitcodes = args['profitcode']

    if not profitcodes:
        subprojectoperatorcode = subprojectoperatorcode or username

    sub_projects = cms.selectSubprojectInfosByCond(
        subprojectoperatorcode=subprojectoperatorcode,
        profitcodes=args['profitcode'])

    logger.info('Total subprojects: \033[1m{}\033[0m [profitcode: {}, subprojectoperatorcode: {}]'.format(
        len(sub_projects), profitcodes, subprojectoperatorcode))

    sub_infos = cms.get_sub_infos(sub_projects, limit=args['limit'], subproject=args['subprojectnum'], project=args['projectnum'])

    logger.info('Total subprojects after filter: \033[1m{total}\033[0m [limit: {limit}, subproject: {subprojectnum}, projectnum: {projectnum}]'.format(
        total=len(sub_infos), **args))

    # exit()

    if args['list']:
        with open(args['out'], 'w') as out:
            for sub_info in sub_infos:
                out.write('{subprojectnum}\n'.format(**sub_info))
    else:
        cms.deal_sub_infos(sub_infos, **args)

    sys.stderr.write('\n')
    logger.info('\033[33msave file: {}\033[0m'.format(args['out']))
    logger.info('\033[32mtime used: {:.1f}s\033[0m'.format(time.time() - start_time))


def parser_add_cms(parser):

    parser.description = __doc__
    parser.epilog = textwrap.dedent('''
        \033[36mexamples:
            %(prog)s -o disease.xlsx
            %(prog)s -o disease.xlsx -r cms_report.xlsx      # 使用CMS导出的报表文件，用于更新结算样本数
            %(prog)s -o disease.xlsx -l 100                  # 只输出100条记录，可用于测试
            %(prog)s -o disease.xlsx -c your_config.xlsx     # 使用指定的配置文件
            %(prog)s -o disease.xlsx -pm Name                # 设置项目经理名字
            %(prog)s -o disease.xlsx -pcode 1911             # 指定部门编号
        \033[0m
    ''')

    group_in_out = parser.add_argument_group('\033[32mInput/Output')
    group_in_out.add_argument('-c', '--config', help='the config excel [default: %(default)s]', default=DEFAULT_CONFIG)
    group_in_out.add_argument('-o', '--out', help='the output filename [default: %(default)s]', default='cms.subproject.xlsx')
    group_in_out.add_argument('-pm', '--PM', help='the name of PM')
    group_in_out.add_argument('-ls', '--list', help='get the subproject list only', action='store_true')
    group_in_out.add_argument('-progress', '--progress', help='show progress during running', action='store_true')

    # external data
    group_external = parser.add_argument_group('\033[33mExternal Data')
    group_external.add_argument('-r', '--report-file', help='the report excel which contains sample informations', nargs='*')
    group_external.add_argument('-lims', '--lims-file', help='the lims data to update stage information')

    # filter argument
    group_filter = parser.add_argument_group('\033[34mFilter Project')
    group_filter.add_argument('-operator', '--subprojectoperatorcode', help='the name of operator manager')
    group_filter.add_argument('-pnum', '--projectnum', help='filter with project list')
    group_filter.add_argument('-spnum', '--subprojectnum', help='filter with subproject list')
    group_filter.add_argument('-pcode', '--profitcode', help='the profitcode to filter', nargs='*')
    group_filter.add_argument('-l', '--limit', help='the limit count of output, just for testing', type=int)

    parser.set_defaults(func=main)
